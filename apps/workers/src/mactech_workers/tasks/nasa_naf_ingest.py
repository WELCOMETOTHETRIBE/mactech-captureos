"""NASA Acquisition Forecast (NAF) — direct XLSX ingest.

Sprint 23. NASA's NAF page is a jQuery DataTables view backed by a
public XLSX at:
  https://www.hq.nasa.gov/office/procurement/forecast/AcqForecastNew.xlsx

~150 rows × 46 columns, refreshed by NASA HQ procurement quarterly.
Bypasses Apify entirely (XLSX is structured + zero auth) and lands
directly into forecasts_raw.
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any

import httpx
import openpyxl
from mactech_db import unscoped_session
from mactech_db.models import ForecastRaw
from sqlalchemy.dialects.postgresql import insert as pg_insert

from mactech_workers.celery_app import celery_app

log = logging.getLogger(__name__)

NAF_XLSX_URL = "https://www.hq.nasa.gov/office/procurement/forecast/AcqForecastNew.xlsx"
NAF_PAGE_URL = "https://www.hq.nasa.gov/office/procurement/forecast/NAF.html"

# NASA's value-range strings → (low, high). They use $-amount banding
# similar to DHS APFS.
_NASA_VALUE_RANGES: dict[str, tuple[Decimal | None, Decimal | None]] = {
    "Less than $250K": (None, Decimal("250000")),
    "$250K - $1M": (Decimal("250000"), Decimal("1000000")),
    "$1M - $5M": (Decimal("1000000"), Decimal("5000000")),
    "$5M - $10M": (Decimal("5000000"), Decimal("10000000")),
    "$10M - $50M": (Decimal("10000000"), Decimal("50000000")),
    "$50M - $100M": (Decimal("50000000"), Decimal("100000000")),
    "$100M - $250M": (Decimal("100000000"), Decimal("250000000")),
    "$100M - $250M ": (Decimal("100000000"), Decimal("250000000")),
    "$100M - $500M": (Decimal("100000000"), Decimal("500000000")),
    "Greater than $500M": (Decimal("500000000"), None),
    "Over $500M": (Decimal("500000000"), None),
}

_QUARTER_TO_END_DATE = {
    "Q1": (12, 31),  # FY Q1 ends Dec 31
    "Q2": (3, 31),
    "Q3": (6, 30),
    "Q4": (9, 30),
}


@dataclass
class NafIngestStats:
    fetched_rows: int
    upserted: int
    skipped: int
    error: str | None
    duration_ms: int


@celery_app.task(name="mactech.nasa_naf.ingest")
def nasa_naf_ingest_task() -> dict[str, Any]:
    return asdict(asyncio.run(_ingest()))


async def _ingest() -> NafIngestStats:
    started = datetime.now(UTC)
    try:
        async with httpx.AsyncClient(
            timeout=httpx.Timeout(60.0, connect=15.0),
            follow_redirects=True,
        ) as client:
            resp = await client.get(NAF_XLSX_URL)
            if resp.status_code != 200:
                return NafIngestStats(
                    fetched_rows=0,
                    upserted=0,
                    skipped=0,
                    error=f"NASA NAF XLSX returned {resp.status_code}",
                    duration_ms=_ms(started),
                )
            xlsx_bytes = resp.content
    except Exception as exc:
        return NafIngestStats(
            fetched_rows=0,
            upserted=0,
            skipped=0,
            error=f"NASA fetch failed: {exc.__class__.__name__}: {exc}",
            duration_ms=_ms(started),
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return NafIngestStats(
            fetched_rows=0,
            upserted=0,
            skipped=0,
            error=f"NASA XLSX parse failed: {exc}",
            duration_ms=_ms(started),
        )

    # NASA file has two sheets — "Forecast Requirements" is the data,
    # "Sheet1" is empty. Look up the right one defensively.
    ws = wb["Forecast Requirements"] if "Forecast Requirements" in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return NafIngestStats(
            fetched_rows=0,
            upserted=0,
            skipped=0,
            error="empty NASA NAF sheet",
            duration_ms=_ms(started),
        )

    # Build a header-name → index map so column-order shifts in NASA's
    # publishing tool don't silently mis-map data.
    idx = {(h or "").strip(): i for i, h in enumerate(header)}

    fetched = 0
    upserted = 0
    skipped = 0
    async with unscoped_session() as session:
        for row in rows_iter:
            fetched += 1
            mapped = _map_row(row, idx)
            if mapped is None:
                skipped += 1
                continue
            stmt = (
                pg_insert(ForecastRaw)
                .values(**mapped)
                .on_conflict_do_update(
                    index_elements=["source_url", "title"],
                    set_={k: v for k, v in mapped.items() if k != "first_seen_at"},
                )
            )
            await session.execute(stmt)
            upserted += 1

    log.info(
        "nasa_naf ingest: rows=%d upserted=%d skipped=%d",
        fetched,
        upserted,
        skipped,
    )
    return NafIngestStats(
        fetched_rows=fetched,
        upserted=upserted,
        skipped=skipped,
        error=None,
        duration_ms=_ms(started),
    )


def _map_row(row: tuple, idx: dict[str, int]) -> dict[str, Any] | None:
    def col(name: str) -> Any:
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    title = _str_or_none(col("TitleOfRequirement"))
    if not title:
        return None

    # Skip already-awarded rows so we don't double-count vs SAM.
    awarded = _str_or_none(col("AwardedOrWithdrawn"))
    if awarded and awarded.lower() == "awarded":
        return None

    naics_raw = _str_or_none(col("NAICS"))
    naics_code = _normalize_naics(naics_raw)

    value_text = _str_or_none(col("EstimatedContractValue"))
    val_low, val_high = _parse_value_range(value_text)

    set_aside = _str_or_none(col("SetAsideType"))
    if set_aside and set_aside.lower() in (
        "no set aside used.",
        "no set aside used",
        "n/a",
        "tbd",
        "to be determined",
    ):
        set_aside = None

    poc_email = _str_or_none(col("TechnicalPOC"))
    if poc_email and "@" not in poc_email:
        poc_email = None
    poc_name = _str_or_none(col("Tech POC Name"))

    sb_email = _str_or_none(col("SmallBusinessSpecialistEmail"))
    if not poc_email and sb_email and "@" in sb_email:
        poc_email = sb_email

    buying_office = _str_or_none(col("BuyingOffice"))
    state = _str_or_none(col("PlaceOfPerformanceState"))
    city = _str_or_none(col("PlaceOfPerformanceCity"))
    contracting_office = buying_office
    if city or state:
        loc = ", ".join(p for p in [city, state] if p and p.lower() != "tbd")
        if loc and contracting_office:
            contracting_office = f"{contracting_office} ({loc})"
        elif loc:
            contracting_office = loc

    expected_release = _resolve_quarter(col("FYofSolOrNOFORelease"), col("QtrSolOrNOFORelease"))
    expected_award = _resolve_quarter(col("AnticipatedFYAward"), col("Anticipated Qtr of Award"))

    description = _str_or_none(col("Description")) or _str_or_none(col("Summary"))

    source_id = _str_or_none(col("SourceID"))
    source_url = f"{NAF_PAGE_URL}#{source_id}" if source_id else NAF_PAGE_URL

    return {
        "source_url": source_url[:2000],
        "source_host": "www.hq.nasa.gov",
        "source_run_id": None,
        "agency": "NASA",
        "contracting_office": contracting_office[:512] if contracting_office else None,
        "title": title[:1000],
        "description": description,
        "naics_code": naics_code,
        "naics_codes": [naics_code] if naics_code else None,
        "set_aside": set_aside,
        "contract_type": _str_or_none(col("ContractType")),
        "estimated_value_low": val_low,
        "estimated_value_high": val_high,
        "estimated_value_text": value_text,
        "expected_solicitation_date": expected_release,
        "expected_award_date": expected_award,
        "period_of_performance_start": None,
        "period_of_performance_end": None,
        "incumbent_name": None,  # NASA doesn't list incumbent here.
        "incumbent_contract_number": None,
        "poc_name": poc_name,
        "poc_email": poc_email,
        "forecast_id": source_id,
        "raw": {
            "buying_office": buying_office,
            "acquisition_status": _str_or_none(col("AcquisitionStatus")),
            "naics_full": naics_raw,
            "psc_code": _str_or_none(col("PSC Code")),
            "psc_desc": _str_or_none(col("PSC Code Description")),
            "directorate": _str_or_none(col("HQMissionDirectorate")),
            "funding_source": _str_or_none(col("FundingSource")),
            "new_or_recompete": _str_or_none(col("NewOrRecompete")),
            "extent_competed": _str_or_none(col("ExtentCompeted")),
            "type_of_award_vehicle": _str_or_none(col("Type of Award/Contract Vehicle")),
            "estimated_value_text": value_text,
        },
        "first_seen_at": datetime.now(UTC),
        "last_seen_at": datetime.now(UTC),
    }


def _str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _normalize_naics(v: Any) -> str | None:
    s = _str_or_none(v)
    if not s:
        return None
    s = s.split(".", 1)[0]
    if s.isdigit() and 4 <= len(s) <= 6:
        return s if len(s) == 6 else s.zfill(6)
    return None


def _parse_value_range(s: str | None) -> tuple[Decimal | None, Decimal | None]:
    if not s:
        return None, None
    bracket = _NASA_VALUE_RANGES.get(s.strip())
    if bracket is not None:
        return bracket
    m = re.search(
        r"\$?([\d.]+)\s*([KMB])\s*[-–—]\s*\$?([\d.]+)\s*([KMB])",
        s,
        re.IGNORECASE,
    )
    if m:
        mult = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}
        try:
            low = Decimal(m.group(1)) * mult[m.group(2).upper()]
            high = Decimal(m.group(3)) * mult[m.group(4).upper()]
            return low, high
        except Exception:
            pass
    return None, None


def _resolve_quarter(fy: Any, quarter: Any) -> date | None:
    """NASA uses (FY, Q) pairs e.g. ('2026', 'Q3') → 2026-06-30 (last
    day of fed Q3). Federal Q1 starts Oct so FY26 Q1 = Oct-Dec 2025."""
    fy_str = _str_or_none(fy)
    q_str = _str_or_none(quarter)
    if not fy_str or not q_str:
        return None
    m = re.match(r"^(\d{4})", fy_str)
    if not m:
        return None
    year = int(m.group(1))
    q_str = q_str.strip().upper()
    qm = re.match(r"^(Q[1-4])", q_str)
    if not qm:
        return None
    end = _QUARTER_TO_END_DATE.get(qm.group(1))
    if not end:
        return None
    cal_year = year - 1 if qm.group(1) == "Q1" else year
    try:
        return date(cal_year, end[0], end[1])
    except ValueError:
        return None


def _ms(started: datetime) -> int:
    return int((datetime.now(UTC) - started).total_seconds() * 1000)
